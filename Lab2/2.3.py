import openpyxl
from random import uniform
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from math import log

align_center=Alignment(horizontal='center',
                       vertical='center',
                       shrink_to_fit=True)


table=[[uniform(1,30) for i in range(20)] for i in range(20)]
rlint=[[0 for i in range(19)] for i in range(19)]
rexpt=[[0 for i in range(19)] for i in range(19)]
e=float(input('Введите процент погрешности взаимосвязи: '))

def exp_value(row):
	a=sum(row)/20
	return a

def dispersion(row):
	a=0
	for i in range(20):
		a+=row[i]**2
	b=a/20-exp_value(row)
	return b

def koeff(rowx, rowy):
	cov=0
	for i in range(20):
		cov+=float(rowx[i])*float(rowy[i])-exp_value(rowx)*exp_value(rowy)
	r=cov/(20*(dispersion(rowx)*dispersion(rowy))**0.5)
	return r

print('Коррелируемые ряды при линейной зависимости:')
for i in range(19):
	for j in range (i+1, 20):
		rlin=koeff(table[i], table[j])
		if abs(rlin*100)>e:
			print(i+1,'-',j+1,'коэф. коррел.= ',rlin)
		rlint[i][j-1]=rlin

print('Коррелируемые ряды при экспоненциальной зависимости:')
for i in range(19):
	for j in range(i+1, 20):
		logrow=[]
		for n in range(20):
			logrow.append(log(table[j][n]))
		rexp=koeff(table[i],logrow)
		if abs(rexp*100)>e:
			print(i+1,'-',j+1,'коэф. коррел.= ',rexp)
		rexpt[i][j-1]=rexp

wb=openpyxl.load_workbook('Lab2.xlsx')

wb.create_sheet('Задание 3.1')
ws=wb['Задание 3.1']
ws.merge_cells('A1:T1')
ws['A1'].value='Таблица случайных значений' 
ws['A1'].alignment=align_center
for row in table:
	ws.append(row)
ws['U1'].value='Матожидание'
ws['V1'].value='Дисперсия'
ws['U1'].alignment=align_center
ws['V1'].alignment=align_center
for row in range(2,22):
	ws.cell(row=row, column=21).value=exp_value(table[row-2])
	ws.cell(row=row, column=22).value=dispersion(table[row-2])

wb.create_sheet('Задание 3.3.1')
ws=wb['Задание 3.3.1']

ws.merge_cells('A1:S1')
ws['A1'].value='Таблица коэффициентов корреляции для линейной зависимости' 
ws['A1'].alignment=align_center
for row in rlint:
    ws.append(row)
wb.create_sheet('Задание 3.3.2')
ws=wb['Задание 3.3.2']

ws.merge_cells('A1:S1')
ws['A1'].value='Таблица коэффициентов корреляции для экспоненциальной зависимости' 
ws['A1'].alignment=align_center
for row in rexpt:
    ws.append(row)

wb.save('Lab2.xlsx')

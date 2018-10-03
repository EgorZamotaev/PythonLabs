import openpyxl
from openpyxl.utils import get_column_letter
from random import uniform, randint

def table():

	for i in range(1, 21):
		for j in range(1, 21):
			val = uniform(1, 30)
			ws.cell(column = j, row = i, value = val)

	val = 0
	i = 0

	while i < 10 :
		col = randint(1, 20)
		row = randint(1, 20)
		ws.cell(column = col, row = row, value = 0)
		i+=1

def vinz():
	for i in range(1,21):
		for j in range(1,21):
			if ws['A1'].value == 0:
				ws.cell(column=1, row=1, value=15)
				continue
			if j==1 and ws[get_column_letter(j) + str(i)].value == 0:
				val = ws[get_column_letter(20) + str(i-1)].value
				ws.cell(column=j, row=i, value=val)
				continue
			if ws[get_column_letter(j) + str(i)].value == 0: 
				val = ws[get_column_letter(j-1) + str(i)].value
				ws.cell(column=j, row=i, value=val)

def lineApr():
	for j in range(1,21):
		sumx = 0 
		sumx2 = 0
		sumy = 0
		sumxy = 0
		for i in range(1,21):
			sumx += i-1
			sumx2 += (i-1)**2
			sumy += ws[get_column_letter(j) + str(i)].value
			sumxy += ws[get_column_letter(j) + str(i)].value*(i-1)
		a = (20*sumxy - sumx*sumy)/(20*sumx2 - sumx**2)
		b = (sumy-a*sumx)/20
		for i in range(1,21):
			if ws[get_column_letter(j) + str(i)].value == 0:
				val = a*i + b
				ws.cell(column=j, row=i, value=val)


def corr():
	n = int(input('Введите номер восстанавливаемого ряда:\n'))
	m = int(input('Введите номер ряда, с которым он коррелирует:\n'))

	for j in range(1, 21):
		if ws[get_column_letter(j) + str(n)].value == 0 and j == 1:
			val = ws[get_column_letter(j + 1) + str(n)].value / (ws[get_column_letter(j) + str(m)].value * ws[get_column_letter(j + 1) + str(m)].value)
			ws.cell(column = j, row = n, value = val)			
		elif ws[get_column_letter(j) + str(n)].value == 0:
			val = ws[get_column_letter(j - 1) + str(n)].value / (ws[get_column_letter(j) + str(m)].value * ws[get_column_letter(j - 1) + str(m)].value)
			ws.cell(column = j, row = n, value = val)	

if __name__ == '__main__':
	wb=openpyxl.load_workbook('Lab2.xlsx')

	#wb.create_sheet('Задание 2')
	ws=wb['Задание 2']
	 #table()
	
	method=int(input('Введите вид алгоритма для восстановления данных:1-Винзорирование,2-Линейная аппроксимация,3-Корреляционное восстановление\n'))

	if method == 1:
		vinz()
	elif method == 2:
		lineApr()
	elif method == 3:
		corr()
	
	wb.save('Lab2.xlsx')

import openpyxl
import matplotlib.pyplot as plt
from random import uniform

coordinates=[]
x=[]
y=[]
apr=[]
sumx=0
sumy=0
sumx2=0
sumxy=0
space=[int(input('Введите минимальное значение координат: ')), int(input('Введите максимальное значение координат: '))]

wb = openpyxl.Workbook()

ws=wb.active

ws.title='Задание1'

coordinates=[[uniform(space[0],space[1]),uniform(space[0],space[1])] for i in range(20)]
print(coordinates)
print(coordinates[0][0])

ws['A1']='x'
ws['B1']='y'
for row in coordinates:
    ws.append(row)
    
for i in range(20):
    x.append(coordinates[i][0])
    y.append(coordinates[i][1])

for i in range(20):
    sumx+=float(x[i])
    sumy+=float(y[i])
    sumx2+=float(x[i])**2
    sumxy+=float(x[i])*float(y[i])
print(sumx,' ',sumy,' ',sumx2,' ',sumxy)

a=(20*sumxy-sumx*sumy)/(20*sumx2-sumx**2)
b=(sumy-a*sumx)/20
for i in range(len(x)):
    apr.append(a*x[i]+b)
        
plt.plot(x,y,'ro',label='Random points')
plt.plot(x,apr,label='Approximation')
plt.xlabel(r'$x$') 
plt.ylabel(r'$y$')
plt.legend(loc='upper left')
plt.grid(True)
plt.show()

wb.save('Lab2.xlsx')

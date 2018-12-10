import openpyxl
import matplotlib.pyplot as plt

table = []
b = [i for i in range(-5, 11, 1)]
a = [b[i] ** 3 for i in range(16)]
c = [3 * a[i] ** 2 for i in range(16)]

for i in range(16):
    table.append([a[i], b[i], c[i]])
    print(table[i])

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Задание 2"

ws['A1'] = "a(b)"
ws['B1'] = "b(c)"
ws['C1'] = "c(a)"
ws['A2'] = "y = x^3"
ws['B2'] = "y = x"
ws['C2'] = "y = x'"

for row in table:
    ws.append(row)

plt.subplots(figsize=(16, 12))
plt.subplot(131)
plt.title("$y = x ^ 3$")
plt.plot(b, a)
plt.subplot(132)
plt.title("$y = x $")
plt.plot(b, b)
plt.subplot(133)
plt.title("$y = \dot x; (y = 3 x ^ 2)$")
plt.plot(a, c)
plt.show()

wb.save('Lab5.xlsx')
